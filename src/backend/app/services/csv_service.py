import csv
import io
import logging
import re
from datetime import datetime
from typing import Any

import openpyxl

from ..models import (
    AssetCreate,
    AssetEnsureRequest,
    AssetProviderSymbolCreate,
    CsvImportCommitResponse,
    CsvImportPreviewResponse,
    CsvImportPreviewRow,
    TransactionCreate,
)
from ..config import get_settings
from ..finance_client import make_finance_client
from ..repository import PortfolioRepository
from ..schemas.instant_portfolio_analyzer import (
    InstantAnalyzeLineError,
    InstantImportedPosition,
    InstantPortfolioImportResponse,
)

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = ["operazione", "isin", "segno", "quantita", "prezzo"]
SEGNO_MAP = {"A": "buy", "V": "sell"}
FEE_COLUMNS = [
    "commissioni fondi sw/ingr/uscita",
    "commissioni fondi banca corrispondente",
    "spese fondi sgr",
    "commissioni amministrato",
]
GENERIC_OPTIONAL_COLUMNS = [
    "titolo",
    "descrizione",
    "divisa",
    "controvalore",
    *FEE_COLUMNS,
]

# Positional column mapping for headerless bank export CSVs (semicolon-separated)
BANK_EXPORT_COLUMNS = [
    "operazione",           # 0  - trade date
    "data_valuta",          # 1  - settlement date
    "descrizione",          # 2  - operation type
    "titolo",               # 3  - instrument name
    "isin",                 # 4  - ISIN code
    "segno",                # 5  - A (buy) / V (sell)
    "quantita",             # 6  - quantity
    "divisa",               # 7  - currency
    "prezzo",               # 8  - price
    "cambio",               # 9  - exchange rate
    "controvalore",         # 10 - total amount
    "commissioni fondi sw/ingr/uscita",          # 11
    "commissioni fondi banca corrispondente",    # 12
    "spese fondi sgr",                           # 13
    "commissioni amministrato",                  # 14
]

FINECO_HEADER_COLUMNS = [
    "operazione",
    "data valuta",
    "descrizione",
    "titolo",
    "isin",
    "segno",
    "quantita",
    "divisa",
    "prezzo",
    "cambio",
    "controvalore",
    "commissioni fondi sw/ingr/uscita",
    "commissioni fondi banca corrispondente",
    "spese fondi sgr",
    "commissioni amministrato",
]

GENERIC_TEMPLATE_COLUMNS = [
    "operazione",
    "isin",
    "segno",
    "quantita",
    "prezzo",
    "titolo",
    "descrizione",
    "divisa",
    "controvalore",
    "commissioni fondi sw/ingr/uscita",
    "commissioni fondi banca corrispondente",
    "spese fondi sgr",
    "commissioni amministrato",
]


BROKER_PROFILES: dict[str, dict[str, Any]] = {
    "fineco": {
        "label": "Fineco",
        "skip_rows": 7,
        "separator": ";",
    },
    "generic": {
        "label": "Generico",
        "skip_rows": 0,
        "separator": None,  # auto-detect
    },
}


def _normalize_header_name(value: str | None) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _split_delimited_line(line: str, delimiter: str) -> list[str]:
    return [_normalize_header_name(part) for part in line.split(delimiter)]


def _is_header_row(cells: list[str], required_columns: list[str]) -> bool:
    normalized = {_normalize_header_name(cell) for cell in cells if _normalize_header_name(cell)}
    required = {_normalize_header_name(column) for column in required_columns}
    return required.issubset(normalized)


def _find_header_row_index(rows: list[list[str]], required_columns: list[str], max_scan_rows: int = 15) -> int | None:
    for index, row in enumerate(rows[:max_scan_rows]):
        if _is_header_row(row, required_columns):
            return index
    return None


def _prepare_fineco_csv_content(file_content: str, skip_rows: int) -> str:
    lines = file_content.splitlines()
    non_empty_lines = [line for line in lines if line.strip()]
    if not non_empty_lines:
        return file_content

    header_index = _find_header_row_index(
        [_split_delimited_line(line, ";") for line in non_empty_lines],
        FINECO_HEADER_COLUMNS,
    )
    if header_index is not None:
        return "\n".join(non_empty_lines[header_index:])

    if skip_rows > 0 and len(non_empty_lines) > skip_rows:
        candidate_lines = non_empty_lines[skip_rows:]
    else:
        candidate_lines = non_empty_lines

    candidate_content = "\n".join(candidate_lines)
    return CsvImportService._detect_and_normalize(candidate_content)


def _read_xlsx_rows(file_bytes: bytes, skip_rows: int = 0, broker: str = "generic") -> list[dict[str, str]]:
    """Read an XLSX file and return rows as list of dicts (same as csv.DictReader)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Il file Excel non contiene fogli")

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    normalized_rows = [[str(cell or "").strip() for cell in row] for row in all_rows]
    required_columns = FINECO_HEADER_COLUMNS if broker == "fineco" else REQUIRED_COLUMNS
    header_index = _find_header_row_index(normalized_rows, required_columns)
    if header_index is None:
        if len(all_rows) <= skip_rows:
            raise ValueError("Il file Excel non contiene righe dati sufficienti")
        header_index = skip_rows

    header_row = all_rows[header_index]
    headers = [str(cell or "").strip() for cell in header_row]
    data_rows = all_rows[header_index + 1:]

    result: list[dict[str, str]] = []
    for row in data_rows:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        row_dict: dict[str, str] = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            val = row[i] if i < len(row) else None
            row_dict[header] = str(val) if val is not None else ""
        result.append(row_dict)

    return result


def _parse_italian_number(s: str) -> float | None:
    """Parse number in Italian or English format.

    Italian: 1.159,77  → 1159.77   (dot = thousands, comma = decimal)
    English: 1,159.77  → 1159.77   (comma = thousands, dot = decimal)
    Plain:   1159.77   → 1159.77
    Plain:   1159,77   → 1159.77

    Heuristic: if both dot and comma are present, the *last* separator is the
    decimal mark.  If only one separator type is present, dots followed by
    exactly 3 digits at the end are thousands separators; otherwise the
    separator is the decimal mark.
    """
    raw = s
    s = s.strip()
    if not s:
        return None

    last_dot = s.rfind(".")
    last_comma = s.rfind(",")

    if last_dot >= 0 and last_comma >= 0:
        if last_comma > last_dot:
            # Italian: 1.234,56
            s = s.replace(".", "").replace(",", ".")
        else:
            # English: 1,234.56
            s = s.replace(",", "")
    elif last_comma >= 0:
        # Only commas – always treat as decimal separator (e.g. 1,0000 → 1.0000)
        s = s.replace(",", ".")
    elif last_dot >= 0:
        # Only dots – thousands separator only if exactly 3 trailing digits
        after_dot = s[last_dot + 1:]
        if len(after_dot) == 3 and after_dot.isdigit() and last_dot != 0:
            s = s.replace(".", "")
        # else: treat dot as decimal (e.g. 1.0000 → 1.0000)

    result = float(s)
    logger.info("_parse_italian_number: raw=%r → normalized=%r → %s", raw.strip(), s, result)
    return result


class CsvImportService:
    def __init__(self, repo: PortfolioRepository) -> None:
        self.repo = repo
        self.settings = get_settings()
        self.finance_client = make_finance_client(self.settings)

    def _resolve_provider_symbol_from_isin(self, isin: str) -> str | None:
        code = (isin or "").strip().upper()
        if not code:
            return None
        try:
            results = self.finance_client.search_symbols(code)
        except Exception:
            return None
        for item in results:
            symbol = (item.symbol or "").strip().upper()
            # Prefer a real ticker over another ISIN-like identifier.
            if symbol and symbol != code and not re.fullmatch(r"[A-Z]{2}[A-Z0-9]{10}", symbol):
                return symbol
        for item in results:
            symbol = (item.symbol or "").strip().upper()
            if symbol and symbol != code:
                return symbol
        return None

    def build_template_xlsx(self, broker: str = "generic") -> tuple[bytes, str]:
        normalized_broker = (broker or "generic").strip().lower()
        if normalized_broker != "generic":
            raise ValueError("Template disponibile solo per il profilo generic")

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        if worksheet is None:
            raise ValueError("Impossibile creare il template Excel")

        worksheet.title = "import_template"
        worksheet.append(GENERIC_TEMPLATE_COLUMNS)

        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        return buffer.getvalue(), "valore365-generic-import-template.xlsx"

    def _ensure_provider_symbol_mapping(self, asset_id: int, provider_symbol: str) -> None:
        value = (provider_symbol or "").strip().upper()
        if not value:
            return
        try:
            self.repo.create_asset_provider_symbol(
                AssetProviderSymbolCreate(
                    asset_id=asset_id,
                    provider=self.settings.finance_provider,
                    provider_symbol=value,
                )
            )
        except ValueError:
            # Mapping may already exist or conflict; don't block import commit.
            return

    @staticmethod
    def _detect_and_normalize(file_content: str) -> str:
        """Detect headerless semicolon-delimited bank exports and prepend a header row."""
        first_line = file_content.split("\n", 1)[0]
        # Heuristic: if the first field looks like a date (dd/mm/yyyy) and the
        # delimiter is semicolon, this is a headerless bank export.
        if ";" in first_line and re.match(r"\d{2}/\d{2}/\d{4}", first_line.strip()):
            header = ";".join(BANK_EXPORT_COLUMNS)
            return header + "\n" + file_content
        return file_content

    def _load_raw_rows(
        self,
        *,
        file_content: str | None,
        filename: str | None,
        file_bytes: bytes | None,
        broker: str,
    ) -> tuple[list[dict[str, str]], list[str]]:
        profile = BROKER_PROFILES.get(broker, BROKER_PROFILES["generic"])
        skip_rows = profile.get("skip_rows", 0)
        is_xlsx = filename and filename.lower().endswith((".xlsx", ".xls"))

        if is_xlsx and file_bytes:
            raw_rows = _read_xlsx_rows(file_bytes, skip_rows=skip_rows, broker=broker)
            if not raw_rows:
                raise ValueError("File Excel vuoto o senza righe dati")
            normalized_fields = [f.strip().lower() for f in raw_rows[0].keys()]
            return raw_rows, normalized_fields

        if file_content:
            if broker == "fineco":
                file_content = _prepare_fineco_csv_content(file_content, skip_rows)
            else:
                if skip_rows > 0:
                    lines = file_content.split("\n")
                    file_content = "\n".join(lines[skip_rows:])
                file_content = self._detect_and_normalize(file_content)
            forced_sep = profile.get("separator")
            delimiter = forced_sep if forced_sep else (";" if ";" in file_content.split("\n", 1)[0] else ",")
            reader = csv.DictReader(io.StringIO(file_content), delimiter=delimiter)
            if reader.fieldnames is None:
                raise ValueError("File CSV vuoto o intestazioni mancanti")
            normalized_fields = [f.strip().lower() for f in reader.fieldnames]
            raw_rows = list(reader)
            return raw_rows, normalized_fields

        raise ValueError("Nessun contenuto file fornito")

    def parse_and_validate(
        self,
        portfolio_id: int,
        user_id: str,
        file_content: str | None = None,
        filename: str | None = None,
        file_bytes: bytes | None = None,
        broker: str = "generic",
    ) -> CsvImportPreviewResponse:
        raw_rows, normalized_fields = self._load_raw_rows(
            file_content=file_content,
            filename=filename,
            file_bytes=file_bytes,
            broker=broker,
        )

        missing = [c for c in REQUIRED_COLUMNS if c not in normalized_fields]
        if missing:
            raise ValueError(f"Colonne obbligatorie mancanti: {', '.join(missing)}")

        rows: list[CsvImportPreviewRow] = []
        valid_count = 0
        error_count = 0

        for idx, raw_row in enumerate(raw_rows, start=1):
            row_data = {k.strip().lower(): (str(v).strip() if v else "") for k, v in raw_row.items()}
            logger.info("Row %d raw_row: %s", idx, dict(raw_row))
            logger.info("Row %d row_data: %s", idx, row_data)
            errors: list[str] = []

            # Parse operazione (date dd/mm/yyyy)
            operazione_str = row_data.get("operazione", "")
            parsed_trade_at: str | None = None
            if not operazione_str:
                errors.append("operazione obbligatorio")
            else:
                try:
                    for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
                        try:
                            dt = datetime.strptime(operazione_str, fmt)
                            parsed_trade_at = dt.isoformat()
                            break
                        except ValueError:
                            continue
                    if parsed_trade_at is None:
                        errors.append(f"Formato data non riconosciuto: {operazione_str}")
                except Exception:
                    errors.append(f"Formato data non valido: {operazione_str}")

            # Parse ISIN
            isin = row_data.get("isin", "").strip().upper()
            if not isin:
                errors.append("isin obbligatorio")

            # Parse segno → side
            segno = row_data.get("segno", "").strip().upper()
            descrizione_lower = row_data.get("descrizione", "").strip().lower()
            # Detect dividends/refunds: blank segno + keyword in description
            if not segno and "dividendo" in descrizione_lower:
                side = "dividend"
            elif not segno and "rimborso" in descrizione_lower:
                side = "dividend"
            else:
                side = SEGNO_MAP.get(segno)
            logger.info("Row %d: segno=%r, descrizione=%r → side=%s", idx, segno, descrizione_lower, side)
            if side is None:
                errors.append(f"segno non valido: '{segno}'. Valori ammessi: A (acquisto), V (vendita)")

            # Parse quantita (Italian number format)
            quantity: float | None = None
            quantita_str = row_data.get("quantita", "")
            if not quantita_str:
                errors.append("quantita obbligatorio")
            else:
                try:
                    quantity = _parse_italian_number(quantita_str)
                    logger.info("Row %d quantita: raw=%r parsed=%s", idx, quantita_str, quantity)
                    if quantity is not None and quantity <= 0:
                        errors.append("quantita deve essere > 0")
                except ValueError:
                    errors.append(f"quantita non numerico: {quantita_str}")

            # Parse prezzo (Italian number format)
            price: float | None = None
            prezzo_str = row_data.get("prezzo", "")
            if not prezzo_str:
                errors.append("prezzo obbligatorio")
            else:
                try:
                    price = _parse_italian_number(prezzo_str)
                    logger.info("Row %d prezzo: raw=%r parsed=%s", idx, prezzo_str, price)
                    if price is not None and price < 0:
                        errors.append("prezzo deve essere >= 0")
                except ValueError:
                    errors.append(f"prezzo non numerico: {prezzo_str}")

            # Use controvalore from CSV to make quantity * price consistent
            # with market price conventions (e.g. bonds quoted as % of nominal).
            # We keep the original quoted price and derive quantity so that
            # quantity * quoted_price = controvalore.  This ensures that
            # quantity * yfinance_close also gives the correct market value.
            controvalore: float | None = None
            controvalore_str = row_data.get("controvalore", "")
            if controvalore_str:
                try:
                    controvalore = _parse_italian_number(controvalore_str)
                except ValueError:
                    errors.append(f"controvalore non numerico: {controvalore_str}")

            if side == "dividend":
                # Dividends: bank export has price=0, real amount in controvalore
                if controvalore is not None and controvalore > 0:
                    quantity = 1.0
                    price = controvalore
            elif (
                controvalore is not None
                and price is not None
                and price > 0
                and controvalore > 0
            ):
                # Derive quantity = controvalore / price so that
                # quantity * price = controvalore.  For normal stocks this
                # is the same as the CSV quantity; for bonds quoted as %
                # (e.g. BOT price=97.60, nominal=10000, cv=9760.75)
                # it becomes 100 instead of 10000, making the math work
                # with market prices from yfinance.
                derived_qty = controvalore / price
                logger.info(
                    "Row %d: controvalore=%s, price=%s → derived_qty=%s (csv_qty=%s)",
                    idx, controvalore, price, derived_qty, quantity,
                )
                quantity = derived_qty

            # Parse divisa → trade_currency
            trade_currency = row_data.get("divisa", "").strip().upper() or None

            # Parse cambio (exchange rate) - stored in notes for reference
            cambio: float | None = None
            cambio_str = row_data.get("cambio", "")
            if cambio_str:
                try:
                    cambio = _parse_italian_number(cambio_str)
                except ValueError:
                    pass  # cambio is optional, skip if unparseable

            # Sum all fee columns
            fees: float = 0.0
            for fee_col in FEE_COLUMNS:
                fee_str = row_data.get(fee_col, "")
                if fee_str:
                    try:
                        fee_val = _parse_italian_number(fee_str)
                        if fee_val is not None:
                            fees += abs(fee_val)
                    except ValueError:
                        errors.append(f"commissione non numerica ({fee_col}): {fee_str}")
            fees_result = fees if fees > 0 else None

            # Build notes from descrizione + titolo + cambio
            titolo = row_data.get("titolo", "").strip() or None
            descrizione = row_data.get("descrizione", "").strip() or None
            notes_parts: list[str] = []
            if descrizione:
                notes_parts.append(descrizione)
            if titolo:
                notes_parts.append(f"Titolo: {titolo}")
            if cambio is not None and cambio != 1.0:
                notes_parts.append(f"Cambio: {cambio}")
            notes = "; ".join(notes_parts) if notes_parts else None

            # Try to resolve asset by ISIN
            asset_id: int | None = None
            asset_name: str | None = None
            if isin and not errors:
                try:
                    matches = self.repo.search_assets(isin)
                    # Match by ISIN field
                    exact = next(
                        (m for m in matches if str(m.get("isin", "")).upper() == isin),
                        None,
                    )
                    if exact:
                        asset_id = int(exact["id"])
                        asset_name = exact.get("name")
                except Exception:
                    pass

            is_valid = len(errors) == 0
            if is_valid:
                valid_count += 1
            else:
                error_count += 1

            rows.append(CsvImportPreviewRow(
                row_number=idx,
                valid=is_valid,
                errors=errors,
                trade_at=parsed_trade_at,
                isin=isin or None,
                titolo=titolo,
                side=side,
                quantity=quantity,
                price=price,
                fees=fees_result,
                taxes=None,
                trade_currency=trade_currency,
                notes=notes,
                asset_id=asset_id,
                asset_name=asset_name,
            ))

        # Save batch to DB
        preview_data = [r.model_dump() for r in rows]
        batch_id = self.repo.create_csv_import_batch(
            portfolio_id=portfolio_id,
            user_id=user_id,
            filename=filename,
            total_rows=len(rows),
            valid_rows=valid_count,
            error_rows=error_count,
            preview_data=preview_data,
        )

        return CsvImportPreviewResponse(
            batch_id=batch_id,
            filename=filename,
            total_rows=len(rows),
            valid_rows=valid_count,
            error_rows=error_count,
            rows=rows,
        )

    def parse_public_portfolio_file(
        self,
        *,
        file_content: str | None = None,
        filename: str | None = None,
        file_bytes: bytes | None = None,
        broker: str = "fineco",
    ) -> InstantPortfolioImportResponse:
        raw_rows, normalized_fields = self._load_raw_rows(
            file_content=file_content,
            filename=filename,
            file_bytes=file_bytes,
            broker=broker,
        )

        missing = [c for c in REQUIRED_COLUMNS if c not in normalized_fields]
        if missing:
            raise ValueError(f"Colonne obbligatorie mancanti: {', '.join(missing)}")

        aggregated_positions: dict[str, dict[str, object]] = {}
        parse_errors: list[InstantAnalyzeLineError] = []
        valid_count = 0
        error_count = 0

        for idx, raw_row in enumerate(raw_rows, start=1):
            row_data = {k.strip().lower(): (str(v).strip() if v else "") for k, v in raw_row.items()}
            errors: list[str] = []

            isin = row_data.get("isin", "").strip().upper()
            if not isin:
                errors.append("isin obbligatorio")

            segno = row_data.get("segno", "").strip().upper()
            descrizione_lower = row_data.get("descrizione", "").strip().lower()
            if not segno and ("dividendo" in descrizione_lower or "rimborso" in descrizione_lower):
                side = "dividend"
            else:
                side = SEGNO_MAP.get(segno)
            if side is None:
                errors.append(f"segno non valido: '{segno}'. Valori ammessi: A (acquisto), V (vendita)")

            quantity: float | None = None
            quantita_str = row_data.get("quantita", "")
            if not quantita_str:
                errors.append("quantita obbligatorio")
            else:
                try:
                    quantity = _parse_italian_number(quantita_str)
                    if quantity is not None and quantity <= 0:
                        errors.append("quantita deve essere > 0")
                except ValueError:
                    errors.append(f"quantita non numerico: {quantita_str}")

            price: float | None = None
            prezzo_str = row_data.get("prezzo", "")
            if not prezzo_str:
                errors.append("prezzo obbligatorio")
            else:
                try:
                    price = _parse_italian_number(prezzo_str)
                    if price is not None and price < 0:
                        errors.append("prezzo deve essere >= 0")
                except ValueError:
                    errors.append(f"prezzo non numerico: {prezzo_str}")

            controvalore: float | None = None
            controvalore_str = row_data.get("controvalore", "")
            if controvalore_str:
                try:
                    controvalore = _parse_italian_number(controvalore_str)
                except ValueError:
                    errors.append(f"controvalore non numerico: {controvalore_str}")

            if errors:
                error_count += 1
                parse_errors.append(
                    InstantAnalyzeLineError(
                        line=idx,
                        raw="; ".join(f"{key}={value}" for key, value in row_data.items() if value),
                        error="; ".join(errors),
                    )
                )
                continue

            valid_count += 1
            if side == "dividend":
                continue

            if controvalore is not None and controvalore > 0:
                exposure_value = abs(controvalore)
            else:
                exposure_value = abs((quantity or 0.0) * (price or 0.0))

            if exposure_value <= 0:
                error_count += 1
                parse_errors.append(
                    InstantAnalyzeLineError(
                        line=idx,
                        raw="; ".join(f"{key}={value}" for key, value in row_data.items() if value),
                        error="Controvalore non disponibile o non valido per questa riga",
                    )
                )
                continue

            signed_value = exposure_value if side == "buy" else -exposure_value
            titolo = row_data.get("titolo", "").strip() or None
            bucket = aggregated_positions.setdefault(
                isin,
                {
                    "identifier": isin,
                    "value": 0.0,
                    "label": titolo,
                    "line": idx,
                },
            )
            bucket["value"] = float(bucket["value"]) + signed_value
            if not bucket.get("label") and titolo:
                bucket["label"] = titolo

        positions = [
            InstantImportedPosition(
                identifier=str(data["identifier"]),
                value=round(float(data["value"]), 2),
                label=str(data["label"]) if data.get("label") else None,
                line=int(data["line"]) if data.get("line") is not None else None,
            )
            for data in aggregated_positions.values()
            if float(data["value"]) > 0
        ]
        positions.sort(key=lambda item: item.value, reverse=True)
        raw_text = "\n".join(f"{position.identifier} {position.value:.2f}" for position in positions)

        return InstantPortfolioImportResponse(
            filename=filename,
            broker=broker,
            total_rows=len(raw_rows),
            valid_rows=valid_count,
            error_rows=error_count,
            positions=positions,
            parse_errors=parse_errors,
            raw_text=raw_text,
        )

    def commit_batch(self, batch_id: int, user_id: str) -> CsvImportCommitResponse:
        batch = self.repo.get_csv_import_batch(batch_id, user_id)
        if batch is None:
            raise ValueError("Batch CSV non trovato")
        if batch["status"] != "pending":
            raise ValueError(f"Batch non in stato pending (stato attuale: {batch['status']})")

        preview_data = batch["preview_data"]
        if not isinstance(preview_data, list):
            raise ValueError("Dati preview non validi")

        portfolio_id = int(batch["portfolio_id"])
        committed = 0
        errors: list[str] = []

        for row_data in preview_data:
            if not row_data.get("valid", False):
                continue

            try:
                # Resolve asset by ISIN
                isin = str(row_data.get("isin", "")).strip().upper()
                asset_id = row_data.get("asset_id")
                side = str(row_data.get("side", "")).lower()

                if asset_id is None and isin:
                    # Try to find asset by ISIN
                    matches = self.repo.search_assets(isin)
                    exact = next(
                        (m for m in matches if str(m.get("isin", "")).upper() == isin),
                        None,
                    )
                    if exact:
                        asset_id = int(exact["id"])
                        provider_symbol = self._resolve_provider_symbol_from_isin(isin)
                        if provider_symbol:
                            self._ensure_provider_symbol_mapping(asset_id, provider_symbol)
                    else:
                        # Auto-create asset. Prefer resolved provider ticker; fallback to ISIN.
                        base_ccy = self.repo.get_portfolio_base_currency(portfolio_id)
                        titolo = row_data.get("titolo") or isin
                        provider_symbol = self._resolve_provider_symbol_from_isin(isin)
                        symbol_for_asset = provider_symbol or isin
                        existing_by_symbol = self.repo.find_asset_by_symbol(symbol_for_asset)
                        if existing_by_symbol:
                            asset_id = int(existing_by_symbol["id"])
                            if provider_symbol:
                                self._ensure_provider_symbol_mapping(asset_id, provider_symbol)
                        else:
                            try:
                                asset = self.repo.create_asset(AssetCreate(
                                    symbol=symbol_for_asset,
                                    name=titolo,
                                    asset_type="stock",
                                    quote_currency=base_ccy,
                                    isin=isin,
                                ))
                                asset_id = asset.id
                                if provider_symbol:
                                    self._ensure_provider_symbol_mapping(asset_id, provider_symbol)
                            except ValueError:
                                # Asset may have been created by a previous row
                                matches = self.repo.search_assets(isin)
                                exact = next(
                                    (m for m in matches if str(m.get("isin", "")).upper() == isin),
                                    None,
                                )
                                if exact:
                                    asset_id = int(exact["id"])
                                    if provider_symbol:
                                        self._ensure_provider_symbol_mapping(asset_id, provider_symbol)
                                else:
                                    existing_by_symbol = self.repo.find_asset_by_symbol(symbol_for_asset)
                                    if existing_by_symbol:
                                        asset_id = int(existing_by_symbol["id"])
                                        if provider_symbol:
                                            self._ensure_provider_symbol_mapping(asset_id, provider_symbol)
                                    else:
                                        errors.append(f"Riga {row_data.get('row_number')}: impossibile risolvere asset ISIN {isin}")
                                        continue

                trade_at_str = row_data.get("trade_at", "")
                trade_at = datetime.fromisoformat(trade_at_str)
                trade_currency = row_data.get("trade_currency") or self.repo.get_portfolio_base_currency(portfolio_id)

                self.repo.create_transaction(
                    TransactionCreate(
                        portfolio_id=portfolio_id,
                        asset_id=asset_id,
                        side=side,
                        trade_at=trade_at,
                        quantity=float(row_data["quantity"]),
                        price=float(row_data["price"]),
                        fees=float(row_data.get("fees") or 0),
                        taxes=float(row_data.get("taxes") or 0),
                        trade_currency=trade_currency,
                        notes=row_data.get("notes"),
                    ),
                    user_id,
                )
                committed += 1
            except Exception as exc:
                errors.append(f"Riga {row_data.get('row_number')}: {exc}")

        self.repo.commit_csv_import_batch(batch_id, user_id)

        return CsvImportCommitResponse(
            batch_id=batch_id,
            committed_transactions=committed,
            errors=errors,
        )
